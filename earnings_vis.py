# ------ earnings_vis.py ------------
#Pull TFSA, RRSP Data from an Excel spreadsheet and visualize contributions, earnings, etc.

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
import seaborn as sns
import pandas as pd
import openpyxl
import mpld3
from mpld3 import plugins
import matplotlib.pyplot as plt
import matplotlib.dates
from datetime import datetime


# TODO: load the file from script argument
# Load the excel file in this dir, extract data on both RRSP,TFSA Sheets
df = openpyxl.load_workbook(r'R-A-Total.xlsx',data_only=True)
TFSA = df['TFSA']
RRSP = df['RRSP']
Margin = df['Margin']

#print(TFSA.max_row)
# Detect actual end of file. TFSA.max_row will give you
# something larger than reality if there are blank cells
#
# Note that the first row is given to column titles, Data starts at row 2

ROW_START = 2
MAX_ROW = ROW_START

for row in range (ROW_START,TFSA.max_row) :
    B_row = 'B' + str(row)
    if TFSA[B_row].value is None:
        break
    else:
        MAX_ROW+=1
#        print(MAX_ROW)
#MAX_ROW+=1
print("Max row: %d" % MAX_ROW)

#  Initialize Datasets for plotting

TFSA_val = []
RRSP_val = []
Margin_val = []
TFSA_cont = []
RRSP_cont= []
Margin_cont= []
Total_val = []
Total_cont = []
Total_earn = []

for row in range(ROW_START, MAX_ROW+1):
    B_row = 'B' + str(row)
    C_row = 'C' + str(row)
    #print('index is: %s ' % index )
    Data1 = TFSA[B_row].value
    Data2 = RRSP[B_row].value
    Data3 = TFSA[C_row].value
    Data4 = RRSP[C_row].value
    Data5 = Margin[B_row].value
    Data6 = Margin[C_row].value
    TFSA_val.append(Data1)
    RRSP_val.append(Data2)
    Margin_val.append(Data5)
    Margin_cont.append(Data6)
    RRSP_cont.append(Data4)
    Total_val.append(Data1+Data2+Data5)
    Total_cont.append(Data3+Data4+Data6)
    Total_earn.append((Data1+Data2+Data5)-(Data3+Data4+Data6))
    #print('MultiSheet Data: %s , %s , %s , %s' % (Data1, Data2,Data3,Data4))

# Make x axis length in years
x=[]
for i in range(MAX_ROW-ROW_START):
    x.append(i/12)

import matplotlib.pyplot as plt
import datetime as dt




print("Min,Max rows are:" + str(ROW_START) + ", " + str(MAX_ROW))
Start_Date = 'A'+str(ROW_START)
End_Date = 'A'+str(MAX_ROW)
print(TFSA[Start_Date].value)
print(TFSA[End_Date].value)




daterange = pd.date_range(TFSA[Start_Date].value,TFSA[End_Date].value,
              freq='MS').strftime("%Y-%m").tolist()
t = [datetime(2018,1,1), datetime(2020,9,30)]
print(daterange)

# Stack Plot Total contributions and total earnings
y = [Total_cont,Total_earn]


fig, plot = plt.subplots(figsize=(6, 6))
#plt.legend(loc='upper left')
sns.set_style(u'whitegrid')
# Change the color and its transparency
#plt.fill_between( x, y1 , color="skyblue", alpha=0.6)
#plt.fill_between( x, y2, color="orange", alpha=0.2)

pal = sns.color_palette("Set2")

print ("Vector Lengths: %d, %d" % (len(y[0]), len(daterange)))
date_form = DateFormatter("%Y")
plot.xaxis.set_major_locator(matplotlib.dates.YearLocator(1))
plot.xaxis.set_major_formatter(matplotlib.dates.DateFormatter('%Y'))
plot.xaxis.set_minor_locator(matplotlib.dates.MonthLocator((1,4,7,10)))

plot.xaxis.set_major_formatter(matplotlib.dates.DateFormatter("\n%Y"))
plot.xaxis.set_minor_formatter(matplotlib.dates.DateFormatter("%b"))

print(daterange[0])
print(daterange[len(daterange)-1])
plot.set_xlim(pd.to_datetime(daterange[0]), pd.to_datetime(daterange[len(daterange)-1]))
plt.setp(plot.get_xticklabels(), rotation=0, ha="center")
plot.xaxis.set_major_formatter(date_form)
#plot.stackplot(daterange,y)
plt.grid()
plt.stackplot(daterange,y,colors=pal, alpha=0.8 )

#plt.tick_params(labelbottom='off')
#plt.tick_params(labelleft='off')
#plt.show()


#plugins.connect(plt.figure())


mpld3.show()




